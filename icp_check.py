# -*- coding: utf-8 -*-
import os
import socks
import socket
import time
import random
import requests
from openpyxl.styles import Alignment
from colorama import Fore, Back, Style, init
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup

# 初始化Colorama，使其在 Windows 终端中正常工作
init(autoreset=True)

global cookies
cookie_str = "__51huid__JfwpT3IBSwA9n8PZ=5a308957-e37b-5e08-ab26-288e5c640b95; __51vcke__JfvlrnUmvss1wiTZ=53a5d85c-39b3-5571-9cb7-8c7066de5572; __51vuft__JfvlrnUmvss1wiTZ=1720192941972; .AspNetCore.Antiforgery.qmZnpHtKIQY=CfDJ8DSDfF06Ji9KnJTVOFYSU3rV-i6hNHBgS3gLebAnFp2lu2Of-4Nj0hZVPuY3NnH5PTFNGd3o6Mz2iPE6vHRBg53sr5haPt-6F7qoNRbZqAfdsnoR-LyG5NHZ1uvlL0ZKmXH1FOtzDuCZ5A8EAPStJaY; __51uvsct__JfvlrnUmvss1wiTZ=2; __vtins__JfvlrnUmvss1wiTZ=%7B%22sid%22%3A%20%2271b845cb-176c-5d8d-89d3-6c4efc5c9bb9%22%2C%20%22vd%22%3A%207%2C%20%22stt%22%3A%20745227%2C%20%22dr%22%3A%2070309%2C%20%22expires%22%3A%201720233172551%2C%20%22ct%22%3A%201720231372551%7D; acw_sc__v2=6688aecd021af1adf590e3c78c9f76190e928fbd; acw_tc=2f624a4e17202348369021640e37329ed42d6792aac4085698bc77f509cdd8"



def parse_cookies(cookie_str):
    cookies = {}
    for item in cookie_str.split('; '):
        key, value = item.split('=', 1)
        cookies[key] = value
    return cookies

cookies = parse_cookies(cookie_str)

#代理设置
# socks.set_default_proxy(socks.SOCKS5, '3.253.57.1', int('51442'), True, 'pSlJvceVAY', 'oNWcAHCZUD')
# socks.set_default_proxy(socks.SOCKS5, '127.0.0.1', int('10808'), True)
# socket.socket = socks.socksocket
proxy_list = [
    (socks.SOCKS5, '127.0.0.1', int(10808), True),
    (socks.SOCKS5, '3.253.57.1', int(51442), True, 'pSlJvceVAY', 'oNWcAHCZUD'),
    (socks.SOCKS5, '3.76.223.185',int(54621),True,'pSlJvceVAY', 'oNWcAHCZUD'),
    (socks.SOCKS5, '18.192.212.82',int(55245),True,'pSlJvceVAY', 'oNWcAHCZUD'),
    (socks.SOCKS5, '34.81.51.21', int(38173),True,'pSlJvceVAY','oNWcAHCZUD'),
    (socks.SOCKS5, '3.120.111.205',int(32644),'pSlJvceVAY','oNWcAHCZUD'),

    # 可以添加更多的代理服务器配置
]

def data_save(html_content):
    if html_content == "":
        pass
    else:
        # 解析HTML内容
        soup = BeautifulSoup(html_content, 'html.parser')

        # 找到目标表格
        table = soup.find('table', {'class': 'table table-sm table-bordered table-hover'})

        # 提取表格数据
        data = []
        for row in table.find_all('tr')[1:]:  # 跳过表头
            cols = row.find_all('td')
            cols = [col.get_text(strip=True) for col in cols]
            # 如果需要处理链接的情况，可以使用下面的代码：
            if row.find('a'):
                cols[1] = row.find('a').get_text(strip=True)
                cols[5] = row.find_all('a')[1].get_text(strip=True)
            data.append(cols)

        # 打印提取的数据
        for item in data:
            print(f"主办单位名称: {item[1]}")
            print(f"主办单位性质: {item[2]}")
            print(f"网站备案号: {item[3]}")
            print(f"网站名称: {item[4]}")
            print(f"网站首页地址: {item[5]}")
            print(f"审核日期: {item[6]}")
            print(f"是否限制接入: {item[7]}")
            print('----')

        # 定义表头
        headers = ["主办单位名称", "主办单位性质", "网站备案号", "网站名称", "网站首页地址", "审核日期", "是否限制接入"]
        # 文件名
        filename = './source/备案信息.xlsx'
        # 浅绿色填充样式
        header_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        # 表头字体样式
        header_font = Font(bold=True)
        # 对齐样式
        alignment = Alignment(horizontal="center", vertical="center")
        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        # 检查文件是否存在
        if os.path.exists(filename):
            workbook = load_workbook(filename)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(headers)
            for col_num, cell in enumerate(sheet[1], 1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = alignment
                cell.border = thin_border

        # 写入数据
        for item in data:
            row_cells = sheet.append(item[1:8])  # 排除掉第一个元素“序号”列
            for row in sheet.iter_rows(min_row=sheet.max_row, max_row=sheet.max_row, min_col=1, max_col=7):
                for cell in row:
                    cell.alignment = alignment
                    cell.border = thin_border

        # 设置列宽
        column_widths = [30, 20, 25, 15, 30, 20, 15]
        for i, width in enumerate(column_widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = width

        # 保存文件
        workbook.save(filename)
        print(f"数据已写入 {filename}")

def read_urls_from_file(file_path):
    try:
        with open(file_path, 'r',encoding="utf-8") as file:
            urls = file.read().splitlines()
        return urls
    except FileNotFoundError:
        print(Fore.RED + Style.BRIGHT + f"File '{file_path}' not found." + Style.RESET_ALL)
        return []

def ask_web_check_ip(top_domain,cookies):
    url = f"https://www.beianx.cn:443/search/{top_domain}"
    headers = {"Cache-Control": "max-age=0", "Upgrade-Insecure-Requests": "1", "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.6167.160 Safari/537.36", "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7", "Sec-Fetch-Site": "same-origin", "Sec-Fetch-Mode": "navigate", "Sec-Fetch-User": "?1", "Sec-Fetch-Dest": "document", "Sec-Ch-Ua": "\"Chromium\";v=\"121\", \"Not A(Brand\";v=\"99\"", "Sec-Ch-Ua-Mobile": "?0", "Sec-Ch-Ua-Platform": "\"Windows\"", "Referer": "https://www.beianx.cn/", "Accept-Encoding": "gzip, deflate, br", "Accept-Language": "zh-CN,zh;q=0.9", "Priority": "u=0, i", "Connection": "close"}
    response = requests.get(url, headers=headers, cookies=cookies).text
    if '没有查询到记录' in response:
        response = ""
        print("暂未查询到备案信息")
    
    return response

def web_check_icp():

    # Specify the file path where URLs are stored
    url_file_path = './source/top_domain.txt'
    urls = read_urls_from_file(url_file_path)

    if not urls:
        print(Fore.RED + Style.NORMAL + "No URLs found in the file. Exiting." + Style.RESET_ALL)
        return

    for url in urls:
        print(Fore.BLUE + Style.NORMAL + url + Style.RESET_ALL)
        sleep_time = random.randrange(0,2)
        time.sleep(sleep_time)
        try:
            html_content = ask_web_check_ip(url,cookies)
            data_save(html_content)
        except Exception as e:
            print(Fore.RED + Style.NORMAL + f"查询失败：{e}" + Style.RESET_ALL)
            pass


def proxy_check():
    print("\n正在检测代理配置~~~~\n\n")

    while proxy_list:
        # 从代理服务器列表中取出一个代理配置
        proxy_config = proxy_list.pop(0)

        # 设置代理服务器
        socks.set_default_proxy(*proxy_config)
        socket.socket = socks.socksocket

        try:
            # 使用代理服务器发起HTTP请求
            response = requests.get('https://www.baidu.com')
            if response.status_code == 200:
                print(Fore.WHITE + Style.NORMAL + f'{proxy_config[1]} 代理服务器存活' + Style.RESET_ALL)
                print(Fore.WHITE + Style.NORMAL + "\n开始查询备案啦\n" + Style.RESET_ALL)
                if web_check_icp():
                    pass
                else:
                    print(Fore.RED + Style.NORMAL + f'查询出错啦：{e}' + Style.RESET_ALL)
            else:
                print(Fore.WHITE + Style.NORMAL + f'{proxy_config[1]} 代理服务器失效，尝试下一个代理' + Style.RESET_ALL)
        except Exception as e:
            print(Fore.WHITE + Style.NORMAL + f'{proxy_config[1]} 代理服务器失效，尝试下一个代理' + Style.RESET_ALL)


if __name__ == '__main__':
    proxy_check()
